import openpyxl

def read_preferences_from_excel(file_path, sheet_name):
    """
    Excelファイルから選好リストを読み込む関数
    :param file_path: Excelファイルのパス
    :param sheet_name: 選好リストが記載されているシート名
    :return: 選好リストを含む辞書
    """
    preferences = {}
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        preferences[row[0]] = [cell for cell in row[1:] if cell is not None]
    return preferences

def read_capacities_from_excel(file_path, sheet_name):
    """
    Excelファイルから受け入れ側の容量を読み込む関数
    :param file_path: Excelファイルのパス
    :param sheet_name: 容量が記載されているシート名
    :return: 容量を含む辞書
    """
    capacities = {}
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        capacities[row[0]] = int(row[1])
    return capacities

def deferred_acceptance(proposers_prefs, acceptors_prefs, acceptors_capacities):
    """
    Deferred Acceptance (DA) アルゴリズムの実装
    :param proposers_prefs: 提案者の選好リスト
    :param acceptors_prefs: 受け入れ側の選好リスト
    :param acceptors_capacities: 受け入れ側の容量
    :return: 最終的なマッチング結果
    """
    proposers = list(proposers_prefs.keys())
    acceptors = list(acceptors_prefs.keys())
    
    # 各受け入れ側の現在のマッチを追跡
    current_matches = {acceptor: [] for acceptor in acceptors}
    
    # 各提案者の次の提案先を追跡
    next_proposal = {proposer: 0 for proposer in proposers}
    
    # マッチしていない提案者を追跡
    unmatched_proposers = set(proposers)
    
    while unmatched_proposers:
        for proposer in list(unmatched_proposers):
            if next_proposal[proposer] < len(proposers_prefs[proposer]):
                acceptor = proposers_prefs[proposer][next_proposal[proposer]]
                next_proposal[proposer] += 1
                
                # 受け入れ側の容量に空きがある場合
                if len(current_matches[acceptor]) < acceptors_capacities[acceptor]:
                    current_matches[acceptor].append(proposer)
                    unmatched_proposers.remove(proposer)
                else:
                    # 現在のマッチと新しい提案者を比較
                    all_candidates = current_matches[acceptor] + [proposer]
                    sorted_candidates = sorted(all_candidates, 
                                               key=lambda x: acceptors_prefs[acceptor].index(x))
                    accepted_candidates = sorted_candidates[:acceptors_capacities[acceptor]]
                    
                    if proposer in accepted_candidates:
                        unmatched_proposers.remove(proposer)
                        rejected_proposer = [p for p in all_candidates if p not in accepted_candidates][0]
                        unmatched_proposers.add(rejected_proposer)
                    
                    current_matches[acceptor] = accepted_candidates
    
    return current_matches

def write_results_to_excel(results, file_path):
    """
    マッチング結果をExcelファイルに書き出す関数
    :param results: マッチング結果の辞書
    :param file_path: 出力するExcelファイルのパス
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Matching Results"
    
    sheet.cell(row=1, column=1, value="Acceptor")
    sheet.cell(row=1, column=2, value="Matched Proposers")
    
    for i, (acceptor, matched_proposers) in enumerate(results.items(), start=2):
        sheet.cell(row=i, column=1, value=acceptor)
        sheet.cell(row=i, column=2, value=", ".join(matched_proposers))
    
    wb.save(file_path)

# メイン処理
if __name__ == "__main__":
    # 入力と出力のファイルパスを設定
    input_file = 'matching_data.xlsx'
    output_file = 'matching_results.xlsx'

    # Excelファイルから選好リストと容量を読み込む
    proposers_prefs = read_preferences_from_excel(input_file, 'Proposers Preferences')
    acceptors_prefs = read_preferences_from_excel(input_file, 'Acceptors Preferences')
    acceptors_capacities = read_capacities_from_excel(input_file, 'Acceptors Capacities')

    # DAアルゴリズムを実行
    result = deferred_acceptance(proposers_prefs, acceptors_prefs, acceptors_capacities)
    print("Matching Results:")
    print(result)

    # 結果をExcelファイルに書き出す
    write_results_to_excel(result, output_file)
    print(f"Results have been written to '{output_file}'")
